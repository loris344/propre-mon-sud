#!/usr/bin/env python3
"""
Convertit Bondash.xlsx -> src/data/seo-pages.json + src/data/internal-links.json

Exécuté UNE SEULE FOIS (ou à chaque mise à jour du plan Excel). Au runtime,
Next.js ne lit plus l'Excel : la source de vérité devient seo-pages.json.

Particularités :
  - Re-calcule les `publishAt` selon un démarrage lent puis accélération
    (décision client) au lieu du 7/jour brut de l'Excel.
  - Marque les pages déjà gérées par une route concrète existante
    (`/` accueil, `/blog/*`) avec handledBy != "catchall" pour que la route
    catch-all ne les génère pas (évite les conflits de routes Next).
  - Ne RÉÉCRIT jamais un MDX : ce script ne touche qu'aux index JSON.

Usage : python3 scripts/excel-to-seo.py
"""
import json
import re
import datetime as dt
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Bondash.xlsx"
OUT_PAGES = ROOT / "src" / "data" / "seo-pages.json"
OUT_LINKS = ROOT / "src" / "data" / "internal-links.json"

# --- Planning de publication (décision client : lent -> rapide) -------------
START_DATE = dt.date(2026, 6, 10)
RAMP_DAYS = 14          # durée de la phase lente
RAMP_PER_DAY = 4        # pages/jour en phase lente
CRUISE_PER_DAY = 7      # pages/jour en vitesse de croisière

# Pages déjà servies par une route concrète existante -> hors catch-all.
EXISTING_ROUTES = {"/", "/blog/"}
def handled_by(url: str) -> str:
    if url in EXISTING_ROUTES:
        return "existing"
    if url.startswith("/blog/"):
        return "blog"          # silo blog traité dans une phase dédiée
    return "catchall"

def split_pipe(v):
    if not v:
        return []
    return [s.strip() for s in str(v).split("|") if s and s.strip()]

def as_bool(v):
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("oui", "true", "1", "yes")

def slug_from_url(url: str) -> str:
    """/desinfection-logement/herault/montpellier/ -> desinfection-logement--herault--montpellier"""
    core = url.strip("/")
    if core == "":
        return "accueil"
    return core.replace("/", "--")

def parse_date(v):
    if not v:
        return None
    if isinstance(v, (dt.date, dt.datetime)):
        return v.date().isoformat() if isinstance(v, dt.datetime) else v.isoformat()
    return str(v)[:10]

# --- Lecture du classeur ----------------------------------------------------
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Pages SEO"]
rows = list(ws.iter_rows(values_only=True))
H = {h: i for i, h in enumerate(rows[0])}
def c(r, name):
    return r[H[name]]

records = []
for r in rows[1:]:
    url = c(r, "URL finale")
    if not url:
        continue
    records.append({
        "id": c(r, "ID") or slug_from_url(url),
        "slug": slug_from_url(url),
        "url": url,
        "title": c(r, "Titre page") or "",
        "type": c(r, "Type de page") or "",
        "template": c(r, "Template utilisé") or "",
        "service": c(r, "Service") or "",
        "departement": c(r, "Département") or "",
        "ville": c(r, "Ville") or "",
        "region": c(r, "Région") or "",
        "priority": c(r, "Priorité"),
        "originalPublishAt": parse_date(c(r, "Date de publication prévue")),
        "publishAt": None,  # recalculé plus bas
        "author": c(r, "Auteur") or "",
        "reviewer": c(r, "Relecteur") or "",
        "keyword": c(r, "Mot-clé principal") or "",
        "secondaryKeywords": split_pipe(c(r, "Mots-clés secondaires")),
        "intent": c(r, "Intention de recherche") or "",
        "parentUrl": c(r, "Page mère URL") or None,
        "childrenUrls": split_pipe(c(r, "Pages enfants URLs")),
        "requiredLinks": split_pipe(c(r, "Liens internes obligatoires")),
        "suggestedLinks": split_pipe(c(r, "Liens internes suggérés")),
        "needsImage": as_bool(c(r, "Besoin image")),
        "imageRequired": as_bool(c(r, "Image obligatoire")),
        "imageType": c(r, "Type image attendue") or "",
        "altProposed": c(r, "Alt image proposé") or "",
        "metaTitle": c(r, "Meta title") or "",
        "metaDescription": c(r, "Meta description") or "",
        "h1": c(r, "H1") or "",
        "h2": split_pipe(c(r, "H2 proposés")),
        "faq": split_pipe(c(r, "FAQ prévue")),
        "schemaType": c(r, "Schema type") or "",
        "canonical": c(r, "Canonical") or "",
        "noindexBeforePublish": as_bool(c(r, "Noindex avant publication")),
        "notes": c(r, "Notes") or "",
        "handledBy": handled_by(url),
    })

# --- Re-calcul du planning (lent -> rapide) ---------------------------------
# On préserve l'ordre de publication voulu par l'Excel (date d'origine puis
# priorité puis ordre de ligne), et on ré-étale selon le rythme client.
def order_key(rec):
    return (rec["originalPublishAt"] or "9999-99-99",
            rec["priority"] if isinstance(rec["priority"], (int, float)) else 999)

sequence = sorted([r for r in records if r["originalPublishAt"]], key=order_key)

def date_for_position(pos: int) -> dt.date:
    """pos = index 0-based dans la séquence de publication."""
    if pos < RAMP_DAYS * RAMP_PER_DAY:
        day_offset = pos // RAMP_PER_DAY
    else:
        rest = pos - RAMP_DAYS * RAMP_PER_DAY
        day_offset = RAMP_DAYS + rest // CRUISE_PER_DAY
    return START_DATE + dt.timedelta(days=day_offset)

for pos, rec in enumerate(sequence):
    rec["publishAt"] = date_for_position(pos).isoformat()

# --- Maillage interne (onglet Maillage) -------------------------------------
wsm = wb["Maillage"]
mrows = list(wsm.iter_rows(values_only=True))
MH = {h: i for i, h in enumerate(mrows[0])}
links = {}
for r in mrows[1:]:
    src = r[MH["Page source"]]
    tgt = r[MH["Page cible"]]
    if not src or not tgt:
        continue
    links.setdefault(src, []).append({
        "target": tgt,
        "anchor": r[MH["Ancre conseillée"]] or "",
        "type": r[MH["Type de lien"]] or "",
        "priority": r[MH["Priorité"]] or "",
    })

# --- Écriture ---------------------------------------------------------------
OUT_PAGES.parent.mkdir(parents=True, exist_ok=True)
meta = {
    "generatedFrom": "Bondash.xlsx",
    "schedule": {
        "startDate": START_DATE.isoformat(),
        "rampDays": RAMP_DAYS, "rampPerDay": RAMP_PER_DAY, "cruisePerDay": CRUISE_PER_DAY,
    },
    "counts": {
        "total": len(records),
        "catchall": sum(1 for r in records if r["handledBy"] == "catchall"),
        "blog": sum(1 for r in records if r["handledBy"] == "blog"),
        "existing": sum(1 for r in records if r["handledBy"] == "existing"),
    },
}
OUT_PAGES.write_text(json.dumps({"meta": meta, "pages": records}, ensure_ascii=False, indent=2) + "\n", "utf-8")
OUT_LINKS.write_text(json.dumps(links, ensure_ascii=False, indent=2) + "\n", "utf-8")

print(f"✓ {OUT_PAGES.relative_to(ROOT)}  ({len(records)} pages)")
print(f"  catch-all={meta['counts']['catchall']}  blog={meta['counts']['blog']}  existing={meta['counts']['existing']}")
print(f"✓ {OUT_LINKS.relative_to(ROOT)}  ({len(links)} sources de liens)")
last = date_for_position(len(sequence) - 1)
print(f"  planning : {START_DATE} -> {last}  ({len(sequence)} pages datées)")
